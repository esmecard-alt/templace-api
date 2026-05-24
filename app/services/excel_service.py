from pathlib import Path
from openpyxl import load_workbook
from app.core.logging import logger


SKIP_SHEETS = {"control", "instrucciones", "readme", "ayuda", "help"}


class ExcelService:

    def _parse_sheet(self, rows: list) -> list | dict:
        if not rows:
            return []

        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]

        data_rows = [row for row in rows[1:] if not all(cell is None for cell in row)]

        if not data_rows:
            return []

        records = []
        for row in data_rows:
            record = {}
            for header, value in zip(headers, row):
                if value is None:
                    record[header] = ""
                elif isinstance(value, float) and value.is_integer():
                    record[header] = int(value)
                elif not isinstance(value, (int, float, bool)):
                    record[header] = str(value)
                else:
                    record[header] = value
            records.append(record)

        # Si solo hay una fila de datos, devuelve objeto en lugar de array
        if len(records) == 1:
            logger.info(f"Hoja con una sola fila — aplanada como objeto")
            return records[0]

        return records

    def excel_to_data(self, excel_path: Path) -> dict:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        data = {}

        for sheet_name in wb.sheetnames:
            if sheet_name.lower() in SKIP_SHEETS:
                logger.info(f"Hoja '{sheet_name}' omitida")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                data[sheet_name] = []
                continue

            parsed = self._parse_sheet(rows)
            data[sheet_name] = parsed

            if isinstance(parsed, dict):
                logger.info(f"Hoja '{sheet_name}' → objeto único")
            else:
                logger.info(f"Hoja '{sheet_name}' → {len(parsed)} fila(s)")

        wb.close()
        logger.info(f"Excel procesado: {list(data.keys())}")
        return data


excel_service = ExcelService()